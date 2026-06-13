"""XLSX/XLSM adapter for workbook documents."""

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from csv_xlsx_editor.domain import CellData, WorkbookDocument, WorksheetDocument


class XlsxAdapter:
    """Loads and saves XLSX/XLSM files while preserving workbook structure."""

    def load(self, path: str | Path) -> WorkbookDocument:
        """Load an XLSX or XLSM file with formulas preserved."""
        file_path = Path(path)
        file_type = "xlsm" if file_path.suffix.lower() == ".xlsm" else "xlsx"
        workbook = load_workbook(
            file_path,
            data_only=False,
            keep_vba=file_type == "xlsm",
        )
        worksheets = [self._worksheet_from_openpyxl(sheet, index) for index, sheet in enumerate(workbook.worksheets)]

        return WorkbookDocument(
            path=file_path,
            file_type=file_type,
            worksheets=worksheets,
            openpyxl_workbook=workbook,
        )

    def save(self, document: WorkbookDocument, path: str | Path) -> None:
        """Save an XLSX/XLSM document back to disk."""
        file_path = Path(path)
        workbook = self.sync_document_to_workbook(document)
        workbook.save(file_path)
        document.path = file_path
        document.mark_clean()

    def sync_document_to_workbook(self, document: WorkbookDocument) -> Workbook:
        """Write document cells into the backing openpyxl workbook."""
        workbook = document.openpyxl_workbook or Workbook()

        for worksheet_document in document.worksheets:
            sheet = self._get_or_create_sheet(workbook, worksheet_document)
            self._sync_worksheet(worksheet_document, sheet)

        return workbook

    def _worksheet_from_openpyxl(self, sheet: Worksheet, index: int) -> WorksheetDocument:
        worksheet = WorksheetDocument(
            sheet_id=f"sheet-{index + 1}",
            title=sheet.title,
            source_sheet_name=sheet.title,
        )

        header_cells: dict[int, CellData] = {}
        headers: list[Any] = []

        if sheet.max_row:
            for cell in sheet[1]:
                column_index = cell.column - 1
                header_value = cell.value
                headers.append(header_value)
                header_cells[column_index] = CellData(
                    value=header_value,
                    formula=None,
                    number_format=cell.number_format,
                    style_id=cell.style_id,
                    readonly_style_ref=copy(cell._style),
                )

        worksheet.set_headers(headers, header_cells=header_cells)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value is None:
                    continue
                row_index = cell.row - 2
                column_index = cell.column - 1
                worksheet.cells[(row_index, column_index)] = CellData(
                    value=cell.value,
                    formula=cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                    number_format=cell.number_format,
                    style_id=cell.style_id,
                    readonly_style_ref=copy(cell._style),
                )
                worksheet.max_row = max(worksheet.max_row, row_index + 1)
                worksheet.max_column = max(worksheet.max_column, column_index + 1)

        worksheet.max_row = max(worksheet.max_row, max(sheet.max_row - 1, 0))
        worksheet.max_column = max(worksheet.max_column, sheet.max_column or 0, len(headers))
        worksheet.rebuild_view()
        return worksheet

    def _get_or_create_sheet(self, workbook: Workbook, worksheet_document: WorksheetDocument) -> Worksheet:
        sheet_name = worksheet_document.source_sheet_name or worksheet_document.title
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

        if len(workbook.worksheets) == 1 and workbook.worksheets[0].title == "Sheet":
            sheet = workbook.worksheets[0]
            sheet.title = sheet_name
            return sheet
        return workbook.create_sheet(sheet_name)

    def _sync_worksheet(self, worksheet_document: WorksheetDocument, sheet: Worksheet) -> None:
        row_order = self._save_row_order(worksheet_document)
        max_columns = max(worksheet_document.max_column, len(worksheet_document.column_headers))

        if worksheet_document.column_headers:
            header_row = 1
            for column_index in range(max_columns):
                target_cell = sheet.cell(row=header_row, column=column_index + 1)
                header_value = (
                    worksheet_document.column_headers[column_index]
                    if column_index < len(worksheet_document.column_headers)
                    else ""
                )
                header_cell = worksheet_document.header_cells.get(column_index)
                self._write_header_cell(target_cell, header_value, header_cell)

        for target_row_index, source_row_index in enumerate(row_order, start=1):
            for source_column_index in range(max_columns):
                target_cell = sheet.cell(row=target_row_index + 1, column=source_column_index + 1)
                cell_data = worksheet_document.get_cell(source_row_index, source_column_index)
                self._write_cell(target_cell, cell_data)

    def _write_cell(
        self,
        target_cell: Any,
        cell_data: CellData,
    ) -> None:
        target_cell.value = cell_data.formula or cell_data.value
        if cell_data.readonly_style_ref is None:
            return
        target_cell._style = copy(cell_data.readonly_style_ref)
        target_cell.number_format = cell_data.number_format

    def _write_header_cell(
        self,
        target_cell: Any,
        value: Any,
        header_cell: CellData | None,
    ) -> None:
        target_cell.value = value
        if header_cell is not None:
            target_cell._style = copy(header_cell.readonly_style_ref)
            target_cell.number_format = header_cell.number_format

    @staticmethod
    def _save_row_order(worksheet: WorksheetDocument) -> list[int]:
        if worksheet.sort_state is None:
            return list(range(worksheet.max_row))

        rows = list(range(worksheet.max_row))
        reverse = worksheet.sort_state.direction == "desc"
        column = worksheet.sort_state.column
        rows.sort(key=lambda row: WorksheetDocument._sort_key_for_cell(worksheet.get_cell(row, column)), reverse=reverse)
        return rows
