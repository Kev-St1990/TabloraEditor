"""Tests for XLSX/XLSM loading and saving."""

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from csv_xlsx_editor.domain import SortState
from csv_xlsx_editor.io import XlsxAdapter


class XlsxAdapterTests(unittest.TestCase):
    """Verify XLSX roundtrips and XLSM load options."""

    def test_load_preserves_formula_as_cell_value(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "formula.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "value"
            sheet["A2"] = 1
            sheet["A3"] = 2
            sheet["A4"] = "=SUM(A2:A3)"
            workbook.save(path)

            document = XlsxAdapter().load(path)

            sheet_document = document.get_active_sheet()
            self.assertEqual(sheet_document.column_headers, ["value"])
            cell = sheet_document.get_cell(2, 0)
            self.assertEqual(cell.value, "=SUM(A2:A3)")
            self.assertEqual(cell.formula, "=SUM(A2:A3)")

    def test_save_preserves_style_when_value_changes(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "styled.xlsx"
            target = Path(temp_dir) / "styled_saved.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "name"
            sheet["A2"] = "old"
            sheet["A2"].font = Font(bold=True, color="00FF0000")
            sheet["A2"].fill = PatternFill("solid", fgColor="0000FF00")
            workbook.save(path)

            adapter = XlsxAdapter()
            document = adapter.load(path)
            document.get_active_sheet().set_cell(0, 0, "new")
            adapter.save(document, target)

            saved = load_workbook(target)
            self.assertEqual(saved.active["A1"].value, "name")
            saved_cell = saved.active["A2"]
            self.assertEqual(saved_cell.value, "new")
            self.assertTrue(saved_cell.font.bold)
            self.assertEqual(saved_cell.fill.fgColor.rgb, "0000FF00")

    def test_save_materializes_sort_order(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "sorted.xlsx"
            target = Path(temp_dir) / "sorted_saved.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["name"])
            sheet.append(["Charlie"])
            sheet.append(["Alice"])
            sheet.append(["Bob"])
            workbook.save(path)

            adapter = XlsxAdapter()
            document = adapter.load(path)
            document.get_active_sheet().apply_sort(SortState(column=0, direction="asc"))
            adapter.save(document, target)

            saved = load_workbook(target)
            self.assertEqual(
                [saved.active.cell(row=row, column=1).value for row in range(1, 5)],
                ["name", "Alice", "Bob", "Charlie"],
            )

    def test_xlsm_load_uses_keep_vba(self) -> None:
        with patch("csv_xlsx_editor.io.xlsx_adapter.load_workbook") as mocked_load:
            workbook = Workbook()
            mocked_load.return_value = workbook

            XlsxAdapter().load("macro.xlsm")

            mocked_load.assert_called_once_with(
                Path("macro.xlsm"),
                data_only=False,
                keep_vba=True,
            )


if __name__ == "__main__":
    unittest.main()
